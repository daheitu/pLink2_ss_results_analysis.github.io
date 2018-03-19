'''
Created on Feb 2, 2018

@author: rqfang
'''

import openpyxl as op
from openpyxl.styles import Font
import csv
import xlrd
import codecs
import os


def get_tag(filename):

    if 'loop-linked' in filename:
        return 'loop'

    elif 'cross-linked' in filename:
        return 'inter'

    elif 'regular' in filename:
        return 'regular'

    else:
        return 'unknown'  #mono with variable modifications (suspended)


def score_compare(evalue_a, evalue_b, score_a, score_b):

    if evalue_a == evalue_b:
        sca = float(score_a)
        scb = float(score_b)

        if sca < scb:
            return True
        else:
            return False
    else:
        eva = float(evalue_a)
        evb = float(evalue_b)

        if eva < evb:
            return True
        else:
            return False


def load_raw_data(filename, data):

    #print(filename)

    file_tag = get_tag(filename)

    infile = open(filename, 'r')
    infiledata = csv.reader(infile, delimiter=',', quotechar='"')

    #===========================================================================
    # wb=op.load_workbook(filename)
    # ws=wb.active
    # ws.title='Result'
    #===========================================================================

    n_rows = 1
    tag = 0
    r_pep = ''
    r_spec = ''
    r_ev = ''
    r_sc = ''
    r_modi = ''
    site = ''
    pepn = ''
    specn = ''

    for line in infiledata:

        if n_rows == 1:
            n_rows += 1
            continue

        if line[0].isdigit():

            if tag != 0:
                data.append([
                    file_tag, site, pepn, specn, r_pep, r_spec, r_ev, r_sc,
                    r_modi
                ])

            tag = 0

            site = line[1]
            pepn = line[2]
            specn = line[3]

            r_pep = ''
            r_spec = ''
            r_ev = ''
            r_sc = ''
            r_modi = ''

        elif line[1].isdigit():

            tag = 1

            t_pep = line[5]
            t_spec = line[2]
            t_ev = line[8]
            t_sc = line[9]
            t_modi = line[7]

            if r_ev == '' or score_compare(t_ev, r_ev, t_sc, r_sc) == True:
                r_pep = t_pep
                r_spec = t_spec
                r_ev = t_ev
                r_sc = t_sc
                r_modi = t_modi

        n_rows += 1

    if tag != 0:
        data.append(
            [file_tag, site, pepn, specn, r_pep, r_spec, r_ev, r_sc, r_modi])

    #print(data)
    return data


def count_cys(modi):

    mods = modi.split(';')

    nmod = []

    n_fixc = 0

    for mod in mods:
        if 'C-1' in mod:
            n_fixc += 1
        else:
            nmod.append(mod)

    if n_fixc == len(mods) or modi == 'null':
        modi = 'null'
    else:
        modi = ';'.join(nmod)

    myres = [modi, n_fixc]

    return myres


def classify(data):

    finalres = {'loop': [], 'inter': [], 'complex': []}

    for info in data:
        
        if info[0] == 'loop':
            t_modi = info[8]

            count_res = count_cys(t_modi)
            t_modi = count_res[0]
            c_num = int(count_res[1])

            if c_num == 2 or c_num == 0:
                finalres[info[0]].append([
                    info[1], info[2], info[3], info[4], info[5], info[6],
                    info[7], t_modi
                ])
            elif c_num % 2 == 0:
                finalres['complex'].append([
                    info[1], info[2], info[3], info[4], info[5], info[6],
                    info[7], t_modi
                ])

        elif info[0] == 'regular':
            t_modi = info[8]

            count_res = count_cys(t_modi)
            t_modi = count_res[0]
            c_num = int(count_res[1])

            if c_num == 2:
                finalres['loop'].append([
                    info[1], info[2], info[3], info[4], info[5], info[6],
                    info[7], t_modi
                ])
            elif c_num != 0 and c_num % 2 == 0:
                finalres['complex'].append([
                    info[1], info[2], info[3], info[4], info[5], info[6],
                    info[7], t_modi
                ])

        elif info[0] == 'inter':
            t_modi = info[8]

            count_res = count_cys(t_modi)
            t_modi = count_res[0]
            c_num = int(count_res[1])

            if c_num == 2 or c_num == 0:
                finalres[info[0]].append([
                    info[1], info[2], info[3], info[4], info[5], info[6],
                    info[7], t_modi
                ])
            elif c_num % 2 == 0:
                finalres['complex'].append([
                    info[1], info[2], info[3], info[4], info[5], info[6],
                    info[7], t_modi
                ])

    return finalres


def output(finalres, filename):

    wb = op.Workbook()
    ws = wb.active

    info_row = [
        'Site_Order', 'Site', 'Unique_Peptide_Number', 'Spectrum_Number',
        'Peptide', 'Spectrum_Title', 'Evalue', 'Score', 'Modifications'
    ]

    mtype = ['loop', 'inter', 'complex']

    row_n = 1

    res_count = 0

    for now_type in mtype:

        res = finalres[now_type]

        ws.cell(row=row_n, column=1).value = now_type.upper()
        ws.cell(row=row_n, column=1).font = Font(bold=True)

        row_n += 1

        for i in range(0, len(info_row)):
            ws.cell(row=row_n, column=i + 1).value = info_row[i]

        row_n += 1

        res_count = 0

        for i in range(0, len(res)):

            res_count += 1
            ws.cell(row=row_n, column=1).value = res_count

            for j in range(0, len(res[i])):
                ws.cell(row=row_n, column=j + 2).value = res[i][j]

            row_n += 1

    wb.save(filename)


def xlsx_to_csv(filename, newfilename):
    workbook = xlrd.open_workbook(filename)
    table = workbook.sheet_by_index(0)
    with codecs.open(newfilename, 'w', encoding='utf-8') as f:
        write = csv.writer(f)
        for row_num in range(table.nrows):
            row_value = table.row_values(row_num)
            write.writerow(row_value)


if __name__ == '__main__':
    print("start ......")
    print("This script was designed for SS_0 workflow")
    foldername = os.getcwd() 
    os.chdir(foldername + "\\reports")
    filelist = os.listdir(os.getcwd())
    link_type_list = [
        "cross-linked_sites", "loop-linked_sites", "mono-linked_sites",
        "regular_proteins"
    ]
    outfilename = foldername + '\\SS_result.xlsx'
    outfilecsv = foldername +  '\\SS_result.csv'
    finalres = []
    data = []
    for fl in filelist:
        link_type = fl[fl.find("filtered_") + 9:fl.find(".csv")]
        if link_type in link_type_list:
            print(fl)
            data = load_raw_data(fl, data)
        else:
            continue
    
    finalres = classify(data)

    output(finalres, outfilename)

    xlsx_to_csv(outfilename, outfilecsv)

    print("The task is done")
